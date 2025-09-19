import os
import sys
import subprocess
import asyncio
import re
import math
import json
import datetime
from collections import defaultdict
from urllib.parse import quote
import msal
import openpyxl
from tqdm import tqdm
import logging

required_modules = ["requests", "msal", "openpyxl", "tqdm", "logging", "aiohttp"]

def install_and_restart():
    try:
        subprocess.check_call([sys.executable.replace('pythonw.exe', 'python.exe'), "-m", "pip", "install"] + required_modules)
        print("Zainstalowano brakujące moduły. Restartowanie skryptu...")
        os.execv(sys.executable.replace('pythonw.exe', 'python.exe'), ['python'] + sys.argv)
    except subprocess.CalledProcessError as e:
        print(f"Błąd podczas instalacji modułów: {e}")
        sys.exit(1)

def check_modules():
    for module in required_modules:
        try:
            __import__(module)
        except ImportError:
            install_and_restart()

check_modules()

import aiohttp


SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILENAME = "email_trend_config.json"
CONFIG_PATH = os.path.join(SCRIPT_DIR, CONFIG_FILENAME)

DEFAULT_CONFIG = {
    "client_id": "",
    "tenant_id": "",
    "client_secret": "",
    "scopes": ["https://graph.microsoft.com/.default"],
    "log_filename": "email_trend_app_only.log",
    "log_level": "INFO",
    "fetch_timeout_seconds": 30,
    "retry_delay_seconds": 5,
    "throttle_delay_seconds": 1,
    "semaphore_limit": 7,
}

REQUIRED_CONFIG_KEYS = ["client_id", "tenant_id", "client_secret"]


def _write_config(config_data):
    try:
        with open(CONFIG_PATH, "w", encoding="utf-8") as config_file:
            json.dump(config_data, config_file, indent=4, ensure_ascii=False)
    except OSError as error:
        print(f"Nie można zapisać pliku konfiguracyjnego {CONFIG_FILENAME}: {error}")
        sys.exit(1)


def load_config():
    if not os.path.exists(CONFIG_PATH):
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as config_file:
                json.dump(DEFAULT_CONFIG, config_file, indent=4, ensure_ascii=False)
        except OSError as error:
            print(
                f"Nie można utworzyć pliku konfiguracyjnego {CONFIG_FILENAME}: {error}"
            )
            sys.exit(1)

        print(
            f"Utworzono plik konfiguracyjny {CONFIG_FILENAME} w lokalizacji {CONFIG_PATH}."
        )
        print("Uzupełnij wymagane dane i uruchom ponownie skrypt.")
        sys.exit(0)

    try:
        with open(CONFIG_PATH, "r", encoding="utf-8") as config_file:
            config_data = json.load(config_file)
    except json.JSONDecodeError as error:
        print(
            f"Błąd analizy pliku konfiguracyjnego {CONFIG_FILENAME}: {error}."
            " Usuń plik lub popraw jego zawartość."
        )
        sys.exit(1)
    except OSError as error:
        print(f"Nie można odczytać pliku konfiguracyjnego {CONFIG_FILENAME}: {error}")
        sys.exit(1)

    if not isinstance(config_data, dict):
        print(
            f"Plik {CONFIG_FILENAME} ma nieprawidłowy format. Oczekiwano obiektu JSON."
        )
        sys.exit(1)

    updated = False
    for key, default_value in DEFAULT_CONFIG.items():
        if key not in config_data:
            config_data[key] = default_value
            updated = True

    if updated:
        try:
            _write_config(config_data)
        except SystemExit:
            raise
        else:
            print(
                "Plik konfiguracyjny został uzupełniony brakującymi ustawieniami."
            )

    missing = [
        key for key in REQUIRED_CONFIG_KEYS if not str(config_data.get(key, "")).strip()
    ]
    if missing:
        missing_values = ", ".join(missing)
        print(
            f"W pliku konfiguracyjnym {CONFIG_FILENAME} brakuje wartości dla: {missing_values}."
        )
        print("Uzupełnij dane i uruchom ponownie skrypt.")
        sys.exit(1)

    return config_data


CONFIG = load_config()


def _read_positive_float(value, default):
    if isinstance(value, bool):
        return default
    try:
        number = float(str(value).strip())
    except (TypeError, ValueError, AttributeError):
        return default
    if math.isnan(number) or math.isinf(number) or number <= 0:
        return default
    return number


def _read_positive_int(value, default):
    number = _read_positive_float(value, default)
    try:
        int_value = int(number)
    except (TypeError, ValueError):
        return default
    if int_value <= 0:
        return default
    return int_value


def _parse_scopes(value):
    if isinstance(value, list):
        scopes = [str(item).strip() for item in value if str(item).strip()]
    elif isinstance(value, str):
        scopes = [item.strip() for item in value.split(",") if item.strip()]
    else:
        scopes = []

    if not scopes:
        scopes = DEFAULT_CONFIG["scopes"]
    return scopes


def _get_float_setting(key):
    raw_value = CONFIG.get(key, DEFAULT_CONFIG[key])
    parsed_value = _read_positive_float(raw_value, DEFAULT_CONFIG[key])
    raw_str = str(raw_value).strip() if raw_value is not None else ""

    if raw_str and parsed_value == DEFAULT_CONFIG[key]:
        try:
            raw_number = float(raw_str)
        except (TypeError, ValueError):
            logging.warning(
                "Nieprawidłowa wartość %s w pliku konfiguracyjnym: %r. Używam domyślnej: %s.",
                key,
                raw_value,
                DEFAULT_CONFIG[key],
            )
        else:
            if math.isnan(raw_number) or math.isinf(raw_number) or raw_number <= 0:
                logging.warning(
                    "Nieprawidłowa wartość %s w pliku konfiguracyjnym: %r. Używam domyślnej: %s.",
                    key,
                    raw_value,
                    DEFAULT_CONFIG[key],
                )

    return parsed_value


def _get_int_setting(key):
    raw_value = CONFIG.get(key, DEFAULT_CONFIG[key])
    parsed_value = _read_positive_int(raw_value, DEFAULT_CONFIG[key])
    raw_str = str(raw_value).strip() if raw_value is not None else ""

    if raw_str and parsed_value == DEFAULT_CONFIG[key]:
        try:
            raw_number = float(raw_str)
        except (TypeError, ValueError):
            logging.warning(
                "Nieprawidłowa wartość %s w pliku konfiguracyjnym: %r. Używam domyślnej: %s.",
                key,
                raw_value,
                DEFAULT_CONFIG[key],
            )
        else:
            if math.isnan(raw_number) or math.isinf(raw_number) or raw_number <= 0:
                logging.warning(
                    "Nieprawidłowa wartość %s w pliku konfiguracyjnym: %r. Używam domyślnej: %s.",
                    key,
                    raw_value,
                    DEFAULT_CONFIG[key],
                )

    return parsed_value


raw_log_filename = str(
    CONFIG.get("log_filename", DEFAULT_CONFIG["log_filename"])
).strip()
LOG_FILENAME = raw_log_filename or DEFAULT_CONFIG["log_filename"]
LOG_FILE_PATH = (
    LOG_FILENAME
    if os.path.isabs(LOG_FILENAME)
    else os.path.join(SCRIPT_DIR, LOG_FILENAME)
)

log_directory = os.path.dirname(LOG_FILE_PATH)
if log_directory:
    os.makedirs(log_directory, exist_ok=True)

raw_log_level = str(
    CONFIG.get("log_level", DEFAULT_CONFIG["log_level"])
).strip().upper()
LOG_LEVEL = getattr(logging, raw_log_level, logging.INFO)

logging.basicConfig(
    level=LOG_LEVEL,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE_PATH, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)

if raw_log_level and raw_log_level != logging.getLevelName(LOG_LEVEL):
    logging.warning(
        "Nieprawidłowa wartość log_level w pliku konfiguracyjnym: %s. Używam %s.",
        raw_log_level,
        logging.getLevelName(LOG_LEVEL),
    )

logging.info("Logi zapisywane do pliku: %s", LOG_FILE_PATH)

CLIENT_ID = CONFIG["client_id"]
TENANT_ID = CONFIG["tenant_id"]
CLIENT_SECRET = CONFIG["client_secret"]
scopes_raw = CONFIG.get("scopes", DEFAULT_CONFIG["scopes"])
SCOPES = _parse_scopes(scopes_raw)

fallback_scopes = False
if isinstance(scopes_raw, list):
    fallback_scopes = not any(str(item).strip() for item in scopes_raw)
elif isinstance(scopes_raw, str):
    fallback_scopes = not scopes_raw.strip()
else:
    fallback_scopes = True

if fallback_scopes:
    logging.warning(
        "Nieprawidłowa wartość scopes w pliku konfiguracyjnym. Używam domyślnej listy: %s.",
        ", ".join(SCOPES),
    )

logging.debug("Zakresy uprawnień używane przez aplikację: %s", ", ".join(SCOPES))

fetch_timeout_seconds = _get_float_setting("fetch_timeout_seconds")
FETCH_TIMEOUT = aiohttp.ClientTimeout(total=fetch_timeout_seconds)

RETRY_DELAY_SECONDS = _get_float_setting("retry_delay_seconds")

THROTTLE_DELAY_SECONDS = _get_float_setting("throttle_delay_seconds")

SEMAPHORE_LIMIT = _get_int_setting("semaphore_limit")
if SEMAPHORE_LIMIT <= 0:
    SEMAPHORE_LIMIT = DEFAULT_CONFIG["semaphore_limit"]

logging.info("Używany plik konfiguracyjny: %s", CONFIG_PATH)
logging.info(
    "Ustawienia żądań: timeout=%ss, retry_delay=%ss, throttle_delay=%ss, limit=%s",
    fetch_timeout_seconds,
    RETRY_DELAY_SECONDS,
    THROTTLE_DELAY_SECONDS,
    SEMAPHORE_LIMIT,
)


def safe_int(value, default=0):
    if isinstance(value, bool):
        return int(value)

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return default
        return int(value)

    if isinstance(value, str):
        cleaned = value.strip()
        if not cleaned:
            return default

        cleaned = (
            cleaned.replace(" ", "")
            .replace("\u00a0", "")
            .replace(",", ".")
        )

        try:
            return int(cleaned)
        except ValueError:
            try:
                return int(float(cleaned))
            except ValueError:
                match = re.search(r"-?\d+", cleaned)
                if match:
                    try:
                        return int(match.group(0))
                    except ValueError:
                        return default
        return default

    return default


def summarize_text(value, max_length=300):
    if value is None:
        return ""

    text = re.sub(r"\s+", " ", str(value)).strip()
    if len(text) > max_length:
        return text[: max_length - 3] + "..."
    return text

def extract_extended_message_size(message):
    for prop in message.get("singleValueExtendedProperties", []) or []:
        prop_id = (prop.get("id") or "").lower()
        if prop_id in {"integer 0x0e08", "long 0x0e08"}:
            size_value = safe_int(prop.get("value"))
            if size_value:
                return size_value
    return 0


def encoded_length(value):
    if not value:
        return 0
    if isinstance(value, bytes):
        return len(value)
    return len(str(value).encode("utf-8", errors="ignore"))


def estimate_message_body_bytes(message):
    total_bytes = 0

    headers = message.get("internetMessageHeaders") or []
    if headers:
        for header in headers:
            name = header.get("name") or ""
            value = header.get("value") or ""
            total_bytes += encoded_length(f"{name}: {value}\r\n")
    else:
        subject = message.get("subject")
        if subject:
            total_bytes += encoded_length(f"Subject: {subject}\r\n")

        sender = message.get("from", {}).get("emailAddress", {}).get("address", "")
        if sender:
            total_bytes += encoded_length(f"From: {sender}\r\n")

        for field in ("toRecipients", "ccRecipients", "bccRecipients"):
            recipients = message.get(field) or []
            if not isinstance(recipients, list):
                continue
            label = field.replace("Recipients", "").upper() or "TO"
            for recipient in recipients:
                address = recipient.get("emailAddress", {}).get("address", "")
                if address:
                    total_bytes += encoded_length(f"{label}: {address}\r\n")

    body = message.get("body") or {}
    body_content = body.get("content") or ""
    total_bytes += encoded_length(body_content)

    if total_bytes == 0:
        total_bytes = encoded_length(message.get("bodyPreview"))

    return total_bytes

def get_access_token():
    authority = f"https://login.microsoftonline.com/{TENANT_ID}"
    app = msal.ConfidentialClientApplication(
        CLIENT_ID,
        authority=authority,
        client_credential=CLIENT_SECRET
    )
    try:
        token_response = app.acquire_token_for_client(scopes=SCOPES)
    except Exception as error:
        logging.exception(
            "Wyjątek podczas uzyskiwania tokena dostępu: %s",
            summarize_text(error),
        )
        raise

    if "access_token" not in token_response:
        error_details = summarize_text(token_response)
        logging.error(
            "Nie udało się uzyskać tokena dostępu dla tenant_id=%s: %s",
            TENANT_ID,
            error_details or "brak szczegółów",
        )
        raise Exception(f"Nie udało się uzyskać tokena: {error_details}")
    return token_response["access_token"]

async def fetch(session, url, headers, semaphore, retries=3, pbar=None):
    attempts_left = retries
    last_error_summary = ""
    async with semaphore:
        while attempts_left > 0:
            try:
                async with session.get(
                    url,
                    headers=headers,
                    timeout=FETCH_TIMEOUT,
                ) as response:
                    if response.status != 200:
                        error_text = await response.text()
                        error_summary = summarize_text(error_text)
                        logging.warning(
                            "Błąd pobierania danych (%s) dla %s: %s",
                            response.status,
                            url,
                            error_summary or "brak treści",
                        )
                        if pbar:
                            pbar.write(
                                f"Błąd pobierania danych: {response.status} {error_text}"
                            )
                        last_error_summary = (
                            f"{response.status} {error_summary}".strip()
                            or last_error_summary
                        )
                    else:
                        return await response.json()
            except (aiohttp.ClientError, asyncio.TimeoutError) as e:
                error_summary = summarize_text(e)
                logging.warning(
                    "Wyjątek podczas pobierania %s: %s",
                    url,
                    error_summary or f"{e.__class__.__name__}: {e}",
                )
                if pbar:
                    pbar.write(
                        f"Błąd połączenia: {e.__class__.__name__}: {e}"
                    )
                last_error_summary = error_summary or f"{e.__class__.__name__}: {e}"

            attempts_left -= 1
            if attempts_left > 0:
                if pbar:
                    pbar.write(
                        f"Ponawianie próby pobierania danych... Pozostało prób: {attempts_left}"
                    )
                logging.debug(
                    "Ponawianie pobierania %s. Pozostało prób: %s",
                    url,
                    attempts_left,
                )
                await asyncio.sleep(RETRY_DELAY_SECONDS)
                await asyncio.sleep(THROTTLE_DELAY_SECONDS)

    if pbar:
        pbar.write("Nie udało się pobrać danych po wielu próbach.")
    logging.error(
        "Nie udało się pobrać danych z %s po %s próbach. Ostatni błąd: %s",
        url,
        retries,
        last_error_summary or "brak szczegółów",
    )
    return None

async def get_child_folders(session, token, mailbox_email, folder, semaphore, path="", pbar=None):
    headers = {"Authorization": f"Bearer {token}"}

    folder_id = folder.get("id")
    folder_name = folder.get("displayName", "")
    total_count = folder.get("totalItemCount", 0)

    if path:
        current_path = f"{path}/{folder_name}"
    else:
        current_path = folder_name

    result = [
        {
            "id": folder_id,
            "path": current_path,
            "displayName": folder_name,
            "totalItemCount": total_count,
        }
    ]

    child_url = f"https://graph.microsoft.com/v1.0/users/{mailbox_email}/mailFolders/{folder_id}/childFolders?$top=100"

    while child_url:
        data = await fetch(session, child_url, headers, semaphore, pbar=pbar)
        if not data:
            logging.warning(
                "Brak danych podfolderów dla %s w ścieżce %s.",
                mailbox_email,
                current_path or folder_name or folder_id,
            )
            break
        children = data.get("value", [])

        for child in children:
            result.extend(await get_child_folders(session, token, mailbox_email, child, semaphore, current_path, pbar))

        child_url = data.get("@odata.nextLink")

    return result

async def get_all_folders(session, token, mailbox_email, semaphore, pbar=None):
    headers = {"Authorization": f"Bearer {token}"}
    url = f"https://graph.microsoft.com/v1.0/users/{mailbox_email}/mailFolders?$top=100"
    all_folders = []

    while url:
        data = await fetch(session, url, headers, semaphore, pbar=pbar)
        if not data:
            logging.error(
                "Nie udało się pobrać listy folderów dla %s (adres żądania: %s).",
                mailbox_email,
                url,
            )
            raise Exception(f"Błąd pobierania folderów dla {mailbox_email}")
        top_folders = data.get("value", [])

        for f in top_folders:
            all_folders.extend(await get_child_folders(session, token, mailbox_email, f, semaphore, path="", pbar=pbar))

        url = data.get("@odata.nextLink")

    return all_folders

async def get_messages_from_folder(session, token, mailbox_email, folder_id, pbar, semaphore, retries=3):
    headers = {
        "Authorization": f"Bearer {token}",
        "Prefer": 'outlook.body-content-type="html"',
    }
    base_url = (
        "https://graph.microsoft.com/v1.0/users/"
        f"{mailbox_email}/mailFolders/{folder_id}/messages"
    )
    attachments_expand = "attachments($select=size,isInline)"
    extended_filter = quote("id eq 'Integer 0x0E08' or id eq 'Long 0x0E08'", safe="")

    select_parts = [
        "subject",
        "receivedDateTime",
        "hasAttachments",
        "id",
        "from",
        "body",
        "bodyPreview",
        "internetMessageHeaders",
        "singleValueExtendedProperties",
        "toRecipients",
        "ccRecipients",
        "bccRecipients",
    ]
    select_clause = ",".join(select_parts)
    expand_clause = ",".join(
        [
            attachments_expand,
            f"singleValueExtendedProperties($filter={extended_filter})",
        ]
    )

    url = f"{base_url}?$select={select_clause}&$expand={expand_clause}&$top=100"
    messages = []

    while url:
        data = await fetch(session, url, headers, semaphore, retries, pbar)
        if not data:
            logging.error(
                "Brak danych wiadomości dla folderu %s w skrzynce %s.",
                folder_id,
                mailbox_email,
            )
            break

        page_messages = data.get("value", [])
        for msg in page_messages:
            attachments = msg.get("attachments", []) or []
            regular_attachment_size = 0
            inline_attachment_size = 0

            for att in attachments:
                att_size = safe_int(att.get("size"))
                if not att_size:
                    continue
                if att.get("isInline"):
                    inline_attachment_size += att_size
                else:
                    regular_attachment_size += att_size

            msg["attachment_size"] = regular_attachment_size

            estimated_body = estimate_message_body_bytes(msg)
            extended_total = extract_extended_message_size(msg)
            baseline_body = estimated_body + inline_attachment_size
            baseline_total = regular_attachment_size + baseline_body

            if extended_total > 0:
                total_size = max(extended_total, baseline_total)
                body_size = max(total_size - regular_attachment_size, baseline_body)
            else:
                body_size = baseline_body
                total_size = baseline_total

            msg["body_size"] = body_size
            msg["total_size"] = total_size

            msg.pop("attachments", None)
            msg.pop("singleValueExtendedProperties", None)
            msg.pop("internetMessageHeaders", None)
            msg.pop("body", None)
            msg.pop("bodyPreview", None)
            msg.pop("toRecipients", None)
            msg.pop("ccRecipients", None)
            msg.pop("bccRecipients", None)

        messages.extend(page_messages)
        pbar.update(len(page_messages))
        url = data.get("@odata.nextLink")

    return messages

def sanitize_sheet_name(name: str) -> str:
    cleaned = re.sub(r'[\\/:\?\*\[\]]+', '_', name)
    return cleaned[:31] or "Folder"

def build_monthly_summary(mailbox_data):
    summary = defaultdict(
        lambda: {
            "message_count": 0,
            "body_size": 0,
            "attachment_size": 0,
            "total_size": 0,
        }
    )
    for folder_path, messages in mailbox_data.items():
        for msg in messages:
            body_bytes = safe_int(msg.get("body_size", 0))
            attachment_bytes = safe_int(msg.get("attachment_size", 0))
            total_bytes = safe_int(
                msg.get("total_size", body_bytes + attachment_bytes)
            )
            received_dt = msg.get("receivedDateTime")
            month_key = "Nieznany"
            if received_dt:
                try:
                    dt_str = received_dt.replace("Z", "")
                    dt_obj = datetime.datetime.fromisoformat(dt_str)
                    month_key = dt_obj.strftime("%Y-%m")
                except ValueError:
                    month_key = received_dt[:7]

            summary[(folder_path, month_key)]["message_count"] += 1
            summary[(folder_path, month_key)]["body_size"] += body_bytes
            summary[(folder_path, month_key)]["attachment_size"] += attachment_bytes
            summary[(folder_path, month_key)]["total_size"] += total_bytes

    return summary

def export_to_excel(data, mailbox_email):
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        default_sheet = wb["Sheet"]
        wb.remove(default_sheet)

    used_sheet_names = set(wb.sheetnames)
    sheet_name_counters = {}

    def get_unique_sheet_name(base_name: str) -> str:
        index = sheet_name_counters.get(base_name, 1)
        while True:
            if index == 1:
                candidate = base_name
            else:
                suffix = f"_{index}"
                allowed_length = max(31 - len(suffix), 0)
                candidate = f"{base_name[:allowed_length]}{suffix}"
            candidate = candidate[:31] or "Folder"
            if candidate not in used_sheet_names:
                used_sheet_names.add(candidate)
                sheet_name_counters[base_name] = index + 1
                return candidate
            index += 1

    for folder_path, messages in data.items():
        base_name = sanitize_sheet_name(folder_path)
        sheet_name = get_unique_sheet_name(base_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append([
            "Subject",
            "Sender",
            "Message Size (bytes)",
            "Message Size (KB)",
            "Message Size (MB)",
            "Attachment Size (bytes)",
            "Attachment Size (KB)",
            "Attachment Size (MB)",
            "Total Size (bytes)",
            "Total Size (KB)",
            "Total Size (MB)",
            "Has Attachments",
            "Received Date",
            "Received Time",
            "Month"
        ])

        for msg in messages:
            subject = msg.get("subject")
            sender = msg.get("from", {}).get("emailAddress", {}).get("address", "")

            body_bytes = safe_int(msg.get("body_size", 0))
            body_kb = round(body_bytes / 1024, 2)
            body_mb = round(body_bytes / (1024 * 1024), 2)

            attach_bytes = safe_int(msg.get("attachment_size", 0))
            attach_kb = round(attach_bytes / 1024, 2)
            attach_mb = round(attach_bytes / (1024 * 1024), 2)

            total_bytes = safe_int(msg.get("total_size", body_bytes + attach_bytes))
            total_kb = round(total_bytes / 1024, 2)
            total_mb = round(total_bytes / (1024 * 1024), 2)

            has_attachments = "Yes" if attach_bytes > 0 else "No"

            received_dt = msg.get("receivedDateTime")
            month_label = ""
            if received_dt:
                try:
                    dt_str = received_dt.replace("Z", "")
                    dt_obj = datetime.datetime.fromisoformat(dt_str)
                    received_date = dt_obj.date().strftime("%Y-%m-%d")
                    received_time = dt_obj.time().strftime("%H:%M:%S")
                    month_label = dt_obj.strftime("%Y-%m")
                except ValueError:
                    received_date = received_dt
                    received_time = ""
                    month_label = received_dt[:7]
            else:
                received_date = ""
                received_time = ""
                month_label = "Nieznany"

            ws.append([
                subject,
                sender,
                body_bytes,
                body_kb,
                body_mb,
                attach_bytes,
                attach_kb,
                attach_mb,
                total_bytes,
                total_kb,
                total_mb,
                has_attachments,
                received_date,
                received_time,
                month_label
            ])

    summary_base_name = "Podsumowanie"
    summary_sheet_name = get_unique_sheet_name(summary_base_name)
    summary_sheet = wb.create_sheet(title=summary_sheet_name)
    summary_sheet.append([
        "Mailbox",
        "Folder",
        "Month",
        "Message Count",
        "Total Size (KB)",
        "Message Size (KB)",
        "Attachment Size (KB)",
        "Total Size (MB)",
        "Message Size (MB)",
        "Attachment Size (MB)"
    ])

    summary_data = build_monthly_summary(data)
    for (folder_path, month_key), values in sorted(summary_data.items(), key=lambda x: (x[0][0], x[0][1])):
        total_size_bytes = safe_int(values.get("total_size", 0))
        body_size_bytes = safe_int(values.get("body_size", 0))
        attachment_size_bytes = safe_int(values.get("attachment_size", 0))
        summary_sheet.append([
            mailbox_email,
            folder_path,
            month_key,
            values["message_count"],
            round(total_size_bytes / 1024, 2),
            round(body_size_bytes / 1024, 2),
            round(attachment_size_bytes / 1024, 2),
            round(total_size_bytes / (1024 * 1024), 2),
            round(body_size_bytes / (1024 * 1024), 2),
            round(attachment_size_bytes / (1024 * 1024), 2)
        ])

    safe_mailbox = mailbox_email.replace("@", "_at_").replace(".", "_")
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{safe_mailbox}_{timestamp}.xlsx"
    wb.save(filename)
    logging.info(f"Dane zapisano do pliku: {filename}")

async def process_mailbox(session, mailbox, token, semaphore):
    try:
        logging.info(f"Przetwarzanie skrzynki: {mailbox}")
        with tqdm(total=1, desc=f"Przetwarzanie {mailbox}", unit="msg", position=0, leave=True) as pbar:
            folders = await get_all_folders(session, token, mailbox, semaphore, pbar)
            total_msgs = sum(f.get("totalItemCount", 0) for f in folders)
            pbar.total = total_msgs

            mailbox_data = {}

            tasks = [
                get_messages_from_folder(
                    session,
                    token,
                    mailbox,
                    f["id"],
                    pbar,
                    semaphore,
                )
                for f in folders
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            for folder_meta, result in zip(folders, results):
                folder_path = folder_meta["path"]
                if isinstance(result, Exception):
                    error_summary = summarize_text(result)
                    logging.warning(
                        "Błąd pobierania folderu %s (%s): %s. Ponawiam próbę...",
                        folder_path,
                        folder_meta.get("id"),
                        error_summary or result.__class__.__name__,
                    )
                    try:
                        retry_messages = await get_messages_from_folder(
                            session,
                            token,
                            mailbox,
                            folder_meta["id"],
                            pbar,
                            semaphore,
                        )
                    except Exception as retry_error:
                        retry_summary = summarize_text(retry_error)
                        logging.error(
                            "Nie udało się pobrać folderu %s po ponownej próbie: %s",
                            folder_path,
                            retry_summary or retry_error.__class__.__name__,
                        )
                        mailbox_data[folder_path] = []
                    else:
                        mailbox_data[folder_path] = retry_messages or []
                    continue

                mailbox_data[folder_path] = result or []

            export_to_excel(mailbox_data, mailbox)
    except Exception:
        logging.exception("Błąd przetwarzania skrzynki %s", mailbox)

async def main():
    logging.info("Rozpoczynam pobieranie danych (app-only)...")
    token = get_access_token()
    logging.info("Token dostępu uzyskany pomyślnie.")

    mailboxes_input = input("Podaj adresy skrzynek oddzielone przecinkiem: ").strip()
    mailbox_list = [m.strip() for m in mailboxes_input.split(",") if m.strip()]

    semaphore = asyncio.Semaphore(SEMAPHORE_LIMIT)

    async with aiohttp.ClientSession() as session:
        tasks = [process_mailbox(session, mailbox, token, semaphore) for mailbox in mailbox_list]
        await asyncio.gather(*tasks)

    logging.info("Przetwarzanie zakończone.")

if __name__ == "__main__":
    asyncio.run(main())
